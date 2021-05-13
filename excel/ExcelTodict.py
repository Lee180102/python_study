# @author   Lee 
# @time     2020/11/1 2:56 下午

import openpyxl


def fileDict(fileList):
    valueDict = {}
    for file_path in fileList:
        workbook = openpyxl.load_workbook(file_path)
        workSheet = workbook['Sheet1']
        for i in range(3, workSheet.max_row+1, 1):
            name = ''
            department = ''
            time = 0
            if workSheet.cell(row=i, column=2).value is not None:
                name = workSheet.cell(row=i, column=2).value
                pass
            if workSheet.cell(row=i, column=3).value is not None:
                department = workSheet.cell(row=i, column=3).value
                pass
            if int(workSheet.cell(row=i, column=8).value):
                time = int(workSheet.cell(row=i, column=8).value)
                pass
            if valueDict.get(name + '_' + department) is None:
                valueDict[name + '_' + department] = time
                pass
            else:
                valueDict[name + '_' + department] = time + int(valueDict.get(name + '_' + department))
                pass
    return valueDict