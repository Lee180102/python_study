import openpyxl

wb = openpyxl.load_workbook('D:\Downloads\demo_example.xlsx', data_only=True)
sheet = wb.active

# metadata_column = {'商业公司': None, '日期': None, '单位名称': None, '商品名称': None, '商品规格': None, '数量': None, '包装单位': None,
#                    '单价': None, '金额': None, '负责人': None}
metadata_column = ('商业公司', '日期', '单位名称', '商品名称', '商品规格', '数量', '包装单位',
                   '单价', '金额', '负责人')

# target_data_column = {
#     'M1-统计年月': None, 'M1-销售客户': None, 'M1-销售日期': None, 'M1-购入客户（导入）': None, 'M1-品种(导入)': None, 'M1-规格(导入)': None,
#     'M1-流向总数量': None, 'M1-含税单价': None, 'M1-流向总金额': None,
#     'M1-备注': None,
#     'M1-购入客户': None, 'M1-产品': None, 'M1-流向级别': None, 'MB-日排序号': None, 'I1-直接负责人': None, 'I1-地区经理': None, 'I1-科室': None,
#     'I1-科室数量': None, 'I1-科室金额': None}

target_data_column = [
    'M1-统计年月', 'M1-销售客户', 'M1-销售日期', 'M1-购入客户（导入）', 'M1-品种(导入)', 'M1-规格(导入)',
    'M1-流向总数量', 'M1-含税单价', 'M1-流向总金额',
    'M1-备注',
    'M1-购入客户', 'M1-产品', 'M1-流向级别', 'MB-日排序号', 'I1-直接负责人', 'I1-地区经理', 'I1-科室',
    'I1-科室数量', 'I1-科室金额']

# {'商业公司': 'A', '日期': 'B', '单位名称': 'C', '商品名称': 'D', '商品规格': 'E', '数量': 'F', '包装单位': 'G', '单价': 'H', '金额': 'I', '负责人': 'J'}
metadata_column_dict = {}

metadata_list = []
startRow = False

for row in sheet.rows:
    for cell in row:
        if cell.value in metadata_column:
            startRow = cell.row + 1
            metadata_column_dict.setdefault(cell.value, openpyxl.utils.get_column_letter(cell.column))


for row in range(startRow, sheet.max_row + 1):
    company_name = sheet[metadata_column_dict.get('单位名称') + str(row)].value
    product_name = sheet[metadata_column_dict.get('商品名称') + str(row)].value
    product_specifications = sheet[metadata_column_dict.get('商品规格') + str(row)].value
    quantity = sheet[metadata_column_dict.get('数量') + str(row)].value

    data_info = (company_name, product_name,
                 product_specifications, quantity)
    metadata_list.append(data_info)

print(metadata_column_dict)
print(metadata_list)
