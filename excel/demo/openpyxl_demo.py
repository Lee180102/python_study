import openpyxl

# wb = openpyxl.load_workbook('D:\Downloads\demo_example.xlsx', data_only=True)

# Getting sheets from the workbook ###################################
# print(wb.sheetnames)
#
# for sheet in wb:
#     print(sheet.title)
#
# mySheet = wb.create_sheet('mySheet')
# print(wb.sheetnames)
#
# sheet2 = wb.get_sheet_by_name('Sheet2')
# print(sheet2)
#
# sheet3 = wb['mySheet']
# print(sheet3)
#
# sheet_active = wb.active
# print(sheet_active)

# Getting cell from the sheets #################################
# sheet_active = wb.active
#
# print(sheet_active)
# print(sheet_active['A1'])
# print(sheet_active['A1'].value)
# c = sheet_active['A1']
#
# print('Row {}, Column {}, Value is {}'.format(c.row, c.column, c.value))
# print('Cell is {}, Value is {}'.format(c.coordinate, c.value))

# print(sheet_active.cell(row=1, column=1))
# print(sheet_active.cell(row=1, column=1).value)
#
# for i in range(1, 8):
#     print(sheet_active.cell(row=i, column=1).value)

# Getting rows and columns from the sheets
# sheet_active = wb.active
# colA = sheet_active['A']
# print(colA[1].value)

# col_range = sheet_active['A:C']

# for col in col_range:
#     for cell in col:
#         print(cell.value)

# row_range = sheet_active[1:6]
#
# for row in row_range:
#     for cell in row:
#         print(cell.value)

# for row in sheet_active.iter_rows(min_row=1, max_row=6, max_col=5):
#     for cell in row:
#         print('Cell is {}, Value is {}'.format(cell.coordinate, cell.value))

# 元组
# print(tuple(sheet_active.rows))

# 切片
# cell_range = sheet_active['A1:I6']
#
# for rowOfCellObjects in cell_range:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate,cellObj.value)
#     print('---------End fo Row------------------')

# 最大行 最大列
# print('{} * {}'.format(sheet_active.max_row, sheet_active.max_column))


# 通过数字获取列的字母 通过字母获取列的数字
# from openpyxl.utils import get_column_letter, column_index_from_string
# print(get_column_letter(2), get_column_letter(47), get_column_letter(900))

# print(column_index_from_string('AHH'))



# Read the spreadsheet data
print('Opening workbook')
wb = openpyxl.load_workbook('D:\Downloads\demo_example.xlsx', data_only=True)

sheet = wb.active


